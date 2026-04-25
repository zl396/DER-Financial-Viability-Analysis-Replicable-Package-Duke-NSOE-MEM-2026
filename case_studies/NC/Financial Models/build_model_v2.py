from pathlib import Path
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HERE = Path(__file__).resolve().parent
wb = Workbook()

# === STYLES ===
blue = Font(name='Arial', color='0000FF', size=10)
blk = Font(name='Arial', color='000000', size=10)
grn = Font(name='Arial', color='008000', size=10)
hdr = Font(name='Arial', bold=True, size=10, color='000000')
sec = Font(name='Arial', bold=True, size=12, color='000000')
ttl = Font(name='Arial', bold=True, size=14, color='000000')
note = Font(name='Arial', italic=True, size=9, color='666666')
yel = PatternFill('solid', fgColor='FFFF00')
lgr = PatternFill('solid', fgColor='F2F2F2')
lbl = PatternFill('solid', fgColor='DCE6F1')
btm = Border(bottom=Side(style='medium'))
dol = '$#,##0;($#,##0);"-"'
dol2 = '$#,##0.00;($#,##0.00);"-"'
pct = '0.0%'
yrfmt = '0'

def sc(ws, r, c, v, f=blk, fmt=None, fill=None):
    cell = ws.cell(row=r, column=c, value=v)
    cell.font = f
    if fmt: cell.number_format = fmt
    if fill: cell.fill = fill
    return cell

# ================================================================
# INPUTS SHEET — Fixed row layout
# ================================================================
ws = wb.active
ws.title = "Inputs"
ws.sheet_properties.tabColor = "4472C4"
ws.column_dimensions['A'].width = 4
ws.column_dimensions['B'].width = 40
ws.column_dimensions['C'].width = 16
ws.column_dimensions['D'].width = 14
ws.column_dimensions['E'].width = 42

sc(ws, 1, 2, "NC Solar Financing Model — Input Parameters", ttl)
sc(ws, 2, 2, "Blue text = editable input  |  Black text = formula  |  Green text = cross-sheet link", note)

# ROW MAP (explicit, matches all formula references):
# C5  = PV+Storage System Cost
# C6  = Solar Only System Cost
# C7  = Solar PV kW
# C8  = Battery kWh
# C10 = (section header)
# C11 = Interest Rate  ← formulas use $C$11
# C12 = Discount Rate  ← formulas use $C$12
# C13 = Loan Term      ← formulas use $C$13
# C14 = Escalation     ← formulas use $C$14
# C16 = (section header)
# C17 = ITC Rate
# C18 = ITC PV+Storage (formula)
# C19 = ITC Solar Only (formula)
# C20 = PowerPair
# C21 = EnergyWise
# C24 = (section header)
# C25 = PV+S RSC without/with
# C26 = PV+S Bridge without/with
# C27 = Solar RSC without/with
# C28 = Solar Bridge without/with

# --- SYSTEM PARAMETERS ---
sc(ws, 4, 2, "SYSTEM PARAMETERS", sec)
sc(ws, 5, 2, "PV + Storage System Cost", blk); sc(ws, 5, 3, 25085, blue, dol, yel); sc(ws, 5, 4, "$", blk)
sc(ws, 5, 5, "Source: SAM model, Duke Energy NC residential config", note)
sc(ws, 6, 2, "Solar Only System Cost", blk); sc(ws, 6, 3, 16302, blue, dol, yel); sc(ws, 6, 4, "$", blk)
sc(ws, 7, 2, "Solar PV Capacity", blk); sc(ws, 7, 3, 5.77, blue, '0.00', yel); sc(ws, 7, 4, "kW", blk)
sc(ws, 8, 2, "Battery Storage Capacity", blk); sc(ws, 8, 3, 13.5, blue, '0.0', yel); sc(ws, 8, 4, "kWh", blk)

# --- FINANCIAL PARAMETERS ---
sc(ws, 10, 2, "FINANCIAL PARAMETERS", sec)
sc(ws, 11, 2, "Debt Interest Rate", blk); sc(ws, 11, 3, 0.0724, blue, pct, yel)
sc(ws, 11, 5, "Source: Prevailing NC residential solar loan rate", note)
sc(ws, 12, 2, "Customer Discount Rate", blk); sc(ws, 12, 3, 0.064, blue, pct, yel)
sc(ws, 13, 2, "Loan Term", blk); sc(ws, 13, 3, 25, blue, yrfmt, yel); sc(ws, 13, 4, "years", blk)
sc(ws, 14, 2, "Electricity Cost Escalation Rate", blk); sc(ws, 14, 3, 0.025, blue, pct, yel)
sc(ws, 14, 5, "Source: EIA historical NC avg ~2-3%/yr", note)

# --- INCENTIVES ---
sc(ws, 16, 2, "INCENTIVES", sec)
sc(ws, 17, 2, "Federal ITC Rate", blk); sc(ws, 17, 3, 0.30, blue, pct, yel)
sc(ws, 17, 5, "Expired Dec 2025; included for historical scenario comparison", note)
sc(ws, 18, 2, "FedITC Amount — PV+Storage", blk); sc(ws, 18, 3, '=C5*C17', blk, dol)
sc(ws, 19, 2, "FedITC Amount — Solar Only", blk); sc(ws, 19, 3, '=C6*C17', blk, dol)
sc(ws, 20, 2, "PowerPair Rebate (PV+Storage only)", blk); sc(ws, 20, 3, 7477.20, blue, dol2, yel)
sc(ws, 20, 5, "Source: Duke Energy NC PowerPair pilot program", note)
sc(ws, 21, 2, "EnergyWise Battery Credit (annual)", blk); sc(ws, 21, 3, 276.51, blue, dol2, yel)
sc(ws, 21, 4, "$/yr", blk)
sc(ws, 21, 5, "Source: Duke Energy EnergyWise Home Battery program", note)

# --- YEAR 1 ELECTRICITY BILLS ---
sc(ws, 23, 2, "YEAR 1 ELECTRICITY BILLS (from SAM simulation)", sec)
sc(ws, 24, 3, "Without System", hdr, None, lgr)
sc(ws, 24, 4, "With System", hdr, None, lgr)
sc(ws, 24, 5, "Annual Savings", hdr, None, lgr)

sc(ws, 25, 2, "PV+Storage — RSC ($/yr)", blk)
sc(ws, 25, 3, 1744, blue, dol, yel); sc(ws, 25, 4, 1113, blue, dol, yel)
sc(ws, 25, 5, '=C25-D25', blk, dol)

sc(ws, 26, 2, "PV+Storage — Bridge ($/yr)", blk)
sc(ws, 26, 3, 1624, blue, dol, yel); sc(ws, 26, 4, 1048, blue, dol, yel)
sc(ws, 26, 5, '=C26-D26', blk, dol)

sc(ws, 27, 2, "Solar Only — RSC ($/yr)", blk)
sc(ws, 27, 3, 1744, blue, dol, yel); sc(ws, 27, 4, 1113, blue, dol, yel)
sc(ws, 27, 5, '=C27-D27', blk, dol)

sc(ws, 28, 2, "Solar Only — Bridge ($/yr)", blk)
sc(ws, 28, 3, 1624, blue, dol, yel); sc(ws, 28, 4, 1121, blue, dol, yel)
sc(ws, 28, 5, '=C28-D28', blk, dol)

print("Inputs sheet done.")

# ================================================================
# CASH FLOW SHEET BUILDER
# ================================================================
def build_cf(ws, sys_label, rate_label, cost_cell, bill_no_cell, bill_w_cell, has_batt, scenarios):
    ws.sheet_properties.tabColor = "70AD47" if has_batt else "ED7D31"
    ws.column_dimensions['A'].width = 30
    for ci in range(2, 30):
        ws.column_dimensions[get_column_letter(ci)].width = 12

    sc(ws, 1, 1, f"{sys_label} — {rate_label} Rate", ttl)
    sc(ws, 2, 1, "All formulas reference Inputs sheet", note)

    row = 4
    net_rows = []

    for si, s in enumerate(scenarios):
        # Scenario header
        sc(ws, row, 1, f"Scenario {si+1}: {s['name']}", sec)
        row += 1

        # Year header
        yr_row = row
        sc(ws, yr_row, 1, "Year", hdr, None, lbl)
        for y in range(26):
            sc(ws, yr_row, y+2, y, hdr, yrfmt, lbl)
        sc(ws, yr_row, 28, "25yr Total", hdr, None, lbl)
        sc(ws, yr_row, 29, "25yr NPV", hdr, None, lbl)
        row += 1

        # --- Debt Balance ---
        r_db = row
        sc(ws, r_db, 1, "Debt Balance ($)", blk)
        # Year 0 initial debt
        parts = [f"Inputs!{cost_cell}"]
        if s['has_itc']:
            itc_cell = "C18" if has_batt else "C19"
            parts.append(f"Inputs!{itc_cell}")
        if s['has_state'] and has_batt:
            parts.append("Inputs!C20")
        yr0_formula = "=" + "-".join(parts)
        ws.cell(row=r_db, column=2, value=yr0_formula).font = grn
        ws.cell(row=r_db, column=2).number_format = dol
        for y in range(1, 26):
            c = y + 2
            p = get_column_letter(c-1)
            ws.cell(row=r_db, column=c, value=f'=MAX(0,{p}{r_db}-{get_column_letter(c)}{r_db+2})').font = blk
            ws.cell(row=r_db, column=c).number_format = dol
        row += 1

        # --- Interest ---
        r_int = row
        sc(ws, r_int, 1, "Interest Payment ($)", blk)
        sc(ws, r_int, 2, 0, blk, dol)
        for y in range(1, 26):
            c = y + 2
            p = get_column_letter(c-1)
            ws.cell(row=r_int, column=c, value=f'=IF({p}{r_db}<=0,0,{p}{r_db}*Inputs!$C$11)').font = blk
            ws.cell(row=r_int, column=c).number_format = dol
        row += 1

        # --- Principal ---
        r_pr = row
        sc(ws, r_pr, 1, "Principal Payment ($)", blk)
        sc(ws, r_pr, 2, 0, blk, dol)
        for y in range(1, 26):
            c = y + 2
            cl = get_column_letter(c)
            ws.cell(row=r_pr, column=c, value=f'=IF(B{r_db}<=0,0,{cl}{r_pr+1}-{cl}{r_int})').font = blk
            ws.cell(row=r_pr, column=c).number_format = dol
        row += 1

        # --- Total P&I ---
        r_pi = row
        sc(ws, r_pi, 1, "Total Payment — P&I ($)", hdr)
        sc(ws, r_pi, 2, 0, blk, dol)
        for y in range(1, 26):
            c = y + 2
            p = get_column_letter(c-1)
            ws.cell(row=r_pi, column=c, value=f'=IF({p}{r_db}<=0,0,ABS(PMT(Inputs!$C$11,Inputs!$C$13,B{r_db},0)))').font = blk
            ws.cell(row=r_pi, column=c).number_format = dol
        ws.cell(row=r_pi, column=28, value=f'=SUM(C{r_pi}:AA{r_pi})').font = blk
        ws.cell(row=r_pi, column=28).number_format = dol
        ws.cell(row=r_pi, column=29, value=f'=NPV(Inputs!$C$12,C{r_pi}:AA{r_pi})').font = blk
        ws.cell(row=r_pi, column=29).number_format = dol
        row += 2  # gap

        # --- Bill without system ---
        r_bn = row
        sc(ws, r_bn, 1, "Elec. Bill — No System ($)", blk)
        sc(ws, r_bn, 2, 0, blk, dol)
        ws.cell(row=r_bn, column=3, value=f'=Inputs!{bill_no_cell}').font = grn
        ws.cell(row=r_bn, column=3).number_format = dol
        for y in range(2, 26):
            c = y + 2
            p = get_column_letter(c-1)
            ws.cell(row=r_bn, column=c, value=f'={p}{r_bn}*(1+Inputs!$C$14)').font = blk
            ws.cell(row=r_bn, column=c).number_format = dol
        row += 1

        # --- Bill with system ---
        r_bw = row
        sc(ws, r_bw, 1, "Elec. Bill — With System ($)", blk)
        sc(ws, r_bw, 2, 0, blk, dol)
        ws.cell(row=r_bw, column=3, value=f'=Inputs!{bill_w_cell}').font = grn
        ws.cell(row=r_bw, column=3).number_format = dol
        for y in range(2, 26):
            c = y + 2
            p = get_column_letter(c-1)
            ws.cell(row=r_bw, column=c, value=f'={p}{r_bw}*(1+Inputs!$C$14)').font = blk
            ws.cell(row=r_bw, column=c).number_format = dol
        row += 1

        # --- Bill Savings ---
        r_bs = row
        sc(ws, r_bs, 1, "Bill Savings ($)", blk)
        sc(ws, r_bs, 2, 0, blk, dol)
        for y in range(1, 26):
            c = y + 2
            cl = get_column_letter(c)
            ws.cell(row=r_bs, column=c, value=f'={cl}{r_bn}-{cl}{r_bw}').font = blk
            ws.cell(row=r_bs, column=c).number_format = dol
        ws.cell(row=r_bs, column=28, value=f'=SUM(C{r_bs}:AA{r_bs})').font = blk
        ws.cell(row=r_bs, column=28).number_format = dol
        ws.cell(row=r_bs, column=29, value=f'=NPV(Inputs!$C$12,C{r_bs}:AA{r_bs})').font = blk
        ws.cell(row=r_bs, column=29).number_format = dol
        row += 1

        # --- Incentive Savings (annual) ---
        r_is = row
        sc(ws, r_is, 1, "Incentive Savings — Annual ($)", blk)
        sc(ws, r_is, 2, 0, blk, dol)
        for y in range(1, 26):
            c = y + 2
            if s['has_state'] and has_batt:
                ws.cell(row=r_is, column=c, value='=Inputs!$C$21').font = grn
            else:
                ws.cell(row=r_is, column=c, value=0).font = blk
            ws.cell(row=r_is, column=c).number_format = dol
        ws.cell(row=r_is, column=28, value=f'=SUM(C{r_is}:AA{r_is})').font = blk
        ws.cell(row=r_is, column=28).number_format = dol
        ws.cell(row=r_is, column=29, value=f'=NPV(Inputs!$C$12,C{r_is}:AA{r_is})').font = blk
        ws.cell(row=r_is, column=29).number_format = dol
        row += 1

        # --- Total Savings ---
        r_ts = row
        sc(ws, r_ts, 1, "Total Savings ($)", hdr)
        for y in range(0, 26):
            c = y + 2
            cl = get_column_letter(c)
            ws.cell(row=r_ts, column=c, value=f'={cl}{r_bs}+{cl}{r_is}').font = blk
            ws.cell(row=r_ts, column=c).number_format = dol
        ws.cell(row=r_ts, column=28, value=f'=SUM(C{r_ts}:AA{r_ts})').font = blk
        ws.cell(row=r_ts, column=28).number_format = dol
        ws.cell(row=r_ts, column=29, value=f'=NPV(Inputs!$C$12,C{r_ts}:AA{r_ts})').font = blk
        ws.cell(row=r_ts, column=29).number_format = dol
        row += 1

        # --- Net Savings ---
        r_ns = row
        sc(ws, r_ns, 1, "Net Savings ($)", Font(name='Arial', bold=True, size=11, color='000000'))
        for y in range(0, 26):
            c = y + 2
            cl = get_column_letter(c)
            ws.cell(row=r_ns, column=c, value=f'={cl}{r_ts}-{cl}{r_pi}').font = blk
            ws.cell(row=r_ns, column=c).number_format = dol
        ws.cell(row=r_ns, column=28, value=f'=SUM(C{r_ns}:AA{r_ns})').font = blk
        ws.cell(row=r_ns, column=28).number_format = dol
        ws.cell(row=r_ns, column=29, value=f'=NPV(Inputs!$C$12,C{r_ns}:AA{r_ns})').font = blk
        ws.cell(row=r_ns, column=29).number_format = dol
        net_rows.append(r_ns)
        
        for c in range(1, 30):
            ws.cell(row=r_ns, column=c).border = btm
        row += 3
    
    return net_rows

# ================================================================
# BUILD ALL SHEETS
# ================================================================
sc_pvs = [
    {"name": "FedITC + State Incentives", "has_itc": True, "has_state": True},
    {"name": "State Incentives Only (No FedITC)", "has_itc": False, "has_state": True},
    {"name": "No Incentives", "has_itc": False, "has_state": False},
]
sc_sol = [
    {"name": "With FedITC", "has_itc": True, "has_state": False},
    {"name": "No Incentives", "has_itc": False, "has_state": False},
]

ws1 = wb.create_sheet("PV+Storage RSC")
nr1 = build_cf(ws1, "PV + Storage", "RSC", "C5", "C25", "D25", True, sc_pvs)
print(f"PV+Storage RSC: net rows at {nr1}")

ws2 = wb.create_sheet("PV+Storage Bridge")
nr2 = build_cf(ws2, "PV + Storage", "Bridge (NMB/TOU)", "C5", "C26", "D26", True, sc_pvs)
print(f"PV+Storage Bridge: net rows at {nr2}")

ws3 = wb.create_sheet("Solar Only RSC")
nr3 = build_cf(ws3, "Solar Only", "RSC", "C6", "C27", "D27", False, sc_sol)
print(f"Solar Only RSC: net rows at {nr3}")

ws4 = wb.create_sheet("Solar Only Bridge")
nr4 = build_cf(ws4, "Solar Only", "Bridge (NMB/TOU)", "C6", "C28", "D28", False, sc_sol)
print(f"Solar Only Bridge: net rows at {nr4}")

# ================================================================
# SUMMARY DASHBOARD
# ================================================================
ws_s = wb.create_sheet("Summary")
ws_s.sheet_properties.tabColor = "7030A0"
ws_s.column_dimensions['A'].width = 14
ws_s.column_dimensions['B'].width = 32
ws_s.column_dimensions['C'].width = 14
ws_s.column_dimensions['D'].width = 15
ws_s.column_dimensions['E'].width = 16
ws_s.column_dimensions['F'].width = 18

sc(ws_s, 1, 1, "NC Solar Financing — Summary Dashboard", ttl)
sc(ws_s, 2, 1, "All values linked to cash flow sheets (green = cross-sheet formula)", note)

headers = ["System", "Scenario", "Rate", "System Cost ($)", "25yr NPV ($)", "Y1 Net Monthly ($)"]
for i, h in enumerate(headers):
    sc(ws_s, 4, i+1, h, hdr, None, lbl)

rows_data = [
    ("PV+Storage", "FedITC + State", "RSC", "'PV+Storage RSC'", nr1[0], "C5"),
    ("PV+Storage", "State Only", "RSC", "'PV+Storage RSC'", nr1[1], "C5"),
    ("PV+Storage", "No Incentives", "RSC", "'PV+Storage RSC'", nr1[2], "C5"),
    ("PV+Storage", "FedITC + State", "Bridge", "'PV+Storage Bridge'", nr2[0], "C5"),
    ("PV+Storage", "State Only", "Bridge", "'PV+Storage Bridge'", nr2[1], "C5"),
    ("PV+Storage", "No Incentives", "Bridge", "'PV+Storage Bridge'", nr2[2], "C5"),
    ("Solar Only", "With FedITC", "RSC", "'Solar Only RSC'", nr3[0], "C6"),
    ("Solar Only", "No Incentives", "RSC", "'Solar Only RSC'", nr3[1], "C6"),
    ("Solar Only", "With FedITC", "Bridge", "'Solar Only Bridge'", nr4[0], "C6"),
    ("Solar Only", "No Incentives", "Bridge", "'Solar Only Bridge'", nr4[1], "C6"),
]

for i, (sys, scn, rate, sheet, net_r, cost_c) in enumerate(rows_data):
    r = 5 + i
    sc(ws_s, r, 1, sys, blk)
    sc(ws_s, r, 2, scn, blk)
    sc(ws_s, r, 3, rate, blk)
    ws_s.cell(row=r, column=4, value=f'=Inputs!{cost_c}').font = grn
    ws_s.cell(row=r, column=4).number_format = dol
    ws_s.cell(row=r, column=5, value=f'={sheet}!AC{net_r}').font = grn
    ws_s.cell(row=r, column=5).number_format = dol
    ws_s.cell(row=r, column=6, value=f'={sheet}!C{net_r}/12').font = grn
    ws_s.cell(row=r, column=6).number_format = dol

print("Summary sheet done.")

# ================================================================
# SAVE
# ================================================================
out = HERE / "NC_Solar_Financing_Model.xlsx"
wb.save(out)
print(f"\nSaved: {out}")
print(f"Sheets: {wb.sheetnames}")
