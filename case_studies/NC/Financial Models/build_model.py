from pathlib import Path
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from copy import copy

HERE = Path(__file__).resolve().parent
wb = Workbook()

# === STYLES ===
blue_font = Font(name='Arial', color='0000FF', size=10)
black_font = Font(name='Arial', color='000000', size=10)
green_font = Font(name='Arial', color='008000', size=10)
header_font = Font(name='Arial', bold=True, size=11, color='000000')
section_font = Font(name='Arial', bold=True, size=12, color='000000')
title_font = Font(name='Arial', bold=True, size=14, color='000000')
yellow_fill = PatternFill('solid', fgColor='FFFF00')
light_gray_fill = PatternFill('solid', fgColor='F2F2F2')
light_blue_fill = PatternFill('solid', fgColor='DCE6F1')
white_fill = PatternFill('solid', fgColor='FFFFFF')
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
bottom_border = Border(bottom=Side(style='medium'))

dollar_fmt = '$#,##0;($#,##0);"-"'
dollar_fmt2 = '$#,##0.00;($#,##0.00);"-"'
pct_fmt = '0.0%'
int_fmt = '#,##0;(#,##0);"-"'
yr_fmt = '0'

def style_cell(ws, row, col, value, font=black_font, fmt=None, fill=None, border=None, align=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font = font
    if fmt: c.number_format = fmt
    if fill: c.fill = fill
    if border: c.border = border
    if align: c.alignment = align
    return c

def style_range(ws, row, col_start, col_end, font=black_font, fmt=None, fill=None):
    for col in range(col_start, col_end + 1):
        c = ws.cell(row=row, column=col)
        c.font = font
        if fmt: c.number_format = fmt
        if fill: c.fill = fill

# ================================================================
# SHEET 1: INPUTS & ASSUMPTIONS
# ================================================================
ws_inp = wb.active
ws_inp.title = "Inputs"
ws_inp.sheet_properties.tabColor = "4472C4"

# Column widths
ws_inp.column_dimensions['A'].width = 4
ws_inp.column_dimensions['B'].width = 38
ws_inp.column_dimensions['C'].width = 16
ws_inp.column_dimensions['D'].width = 8
ws_inp.column_dimensions['E'].width = 40

# Title
style_cell(ws_inp, 1, 1, "NC Solar Financing Model — Input Parameters", title_font)
style_cell(ws_inp, 2, 1, None)
style_cell(ws_inp, 2, 2, "Blue = editable input  |  Black = formula  |  Green = cross-sheet link", Font(name='Arial', italic=True, size=9, color='666666'))

# --- SYSTEM PARAMETERS ---
r = 4
style_cell(ws_inp, r, 2, "SYSTEM PARAMETERS", section_font)
r += 1
# Solar+Storage
style_cell(ws_inp, r, 2, "PV + Storage System Cost", black_font)
style_cell(ws_inp, r, 3, 25085, blue_font, dollar_fmt, yellow_fill)
style_cell(ws_inp, r, 4, "$", black_font)
style_cell(ws_inp, r, 5, "Source: SAM model, Duke Energy NC residential", Font(name='Arial', italic=True, size=9, color='666666'))
r += 1
style_cell(ws_inp, r, 2, "Solar Only System Cost", black_font)
style_cell(ws_inp, r, 3, 16302, blue_font, dollar_fmt, yellow_fill)
style_cell(ws_inp, r, 4, "$", black_font)
r += 1
style_cell(ws_inp, r, 2, "Solar PV Capacity", black_font)
style_cell(ws_inp, r, 3, 5.77, blue_font, '0.00', yellow_fill)
style_cell(ws_inp, r, 4, "kW", black_font)
r += 1
style_cell(ws_inp, r, 2, "Battery Storage Capacity", black_font)
style_cell(ws_inp, r, 3, 13.5, blue_font, '0.0', yellow_fill)
style_cell(ws_inp, r, 4, "kWh", black_font)

# --- FINANCIAL PARAMETERS ---
r += 2  # r=11
style_cell(ws_inp, r, 2, "FINANCIAL PARAMETERS", section_font)
r += 1  # r=12
style_cell(ws_inp, r, 2, "Debt Interest Rate", black_font)
style_cell(ws_inp, r, 3, 0.0724, blue_font, pct_fmt, yellow_fill)
r += 1
style_cell(ws_inp, r, 2, "Customer Discount Rate", black_font)
style_cell(ws_inp, r, 3, 0.064, blue_font, pct_fmt, yellow_fill)
r += 1
style_cell(ws_inp, r, 2, "Loan Term", black_font)
style_cell(ws_inp, r, 3, 25, blue_font, yr_fmt, yellow_fill)
style_cell(ws_inp, r, 4, "years", black_font)
r += 1
style_cell(ws_inp, r, 2, "Electricity Cost Escalation Rate", black_font)
style_cell(ws_inp, r, 3, 0.025, blue_font, pct_fmt, yellow_fill)
style_cell(ws_inp, r, 5, "Source: EIA historical NC avg ~2-3%/yr", Font(name='Arial', italic=True, size=9, color='666666'))

# --- INCENTIVES ---
r += 2  # r=18
style_cell(ws_inp, r, 2, "INCENTIVES", section_font)
r += 1  # r=19
style_cell(ws_inp, r, 2, "Federal ITC Rate", black_font)
style_cell(ws_inp, r, 3, 0.30, blue_font, pct_fmt, yellow_fill)
style_cell(ws_inp, r, 5, "Expired Dec 2025; included for historical scenarios", Font(name='Arial', italic=True, size=9, color='666666'))
r += 1
style_cell(ws_inp, r, 2, "FedITC — PV+Storage", black_font)
style_cell(ws_inp, r, 3, '=C5*C19', black_font, dollar_fmt)
r += 1
style_cell(ws_inp, r, 2, "FedITC — Solar Only", black_font)
style_cell(ws_inp, r, 3, '=C6*C19', black_font, dollar_fmt)
r += 1
style_cell(ws_inp, r, 2, "PowerPair Rebate (PV+Storage only)", black_font)
style_cell(ws_inp, r, 3, 7477.20, blue_font, dollar_fmt2, yellow_fill)
style_cell(ws_inp, r, 5, "Source: Duke Energy NC PowerPair pilot", Font(name='Arial', italic=True, size=9, color='666666'))
r += 1
style_cell(ws_inp, r, 2, "EnergyWise Battery Credit (annual)", black_font)
style_cell(ws_inp, r, 3, 276.51, blue_font, dollar_fmt2, yellow_fill)
style_cell(ws_inp, r, 4, "$/yr", black_font)
style_cell(ws_inp, r, 5, "Source: Duke Energy EnergyWise Home program", Font(name='Arial', italic=True, size=9, color='666666'))

# --- ELECTRICITY BILLS (Year 1) ---
r += 2  # r=26
style_cell(ws_inp, r, 2, "YEAR 1 ELECTRICITY BILLS (from SAM)", section_font)
r += 1
# Headers
style_cell(ws_inp, r, 3, "Without System", header_font, None, light_gray_fill)
style_cell(ws_inp, r, 4, "With System", header_font, None, light_gray_fill)
style_cell(ws_inp, r, 5, "Savings", header_font, None, light_gray_fill)
# Need wider columns for this section
ws_inp.column_dimensions['D'].width = 14
r += 1  # r=28
style_cell(ws_inp, r, 2, "PV+Storage — RSC ($/yr)", black_font)
style_cell(ws_inp, r, 3, 1744, blue_font, dollar_fmt, yellow_fill)
style_cell(ws_inp, r, 4, 1113, blue_font, dollar_fmt, yellow_fill)
style_cell(ws_inp, r, 5, '=C28-D28', black_font, dollar_fmt)
r += 1
style_cell(ws_inp, r, 2, "PV+Storage — Bridge ($/yr)", black_font)
style_cell(ws_inp, r, 3, 1624, blue_font, dollar_fmt, yellow_fill)
style_cell(ws_inp, r, 4, 1048, blue_font, dollar_fmt, yellow_fill)
style_cell(ws_inp, r, 5, '=C29-D29', black_font, dollar_fmt)
r += 1
style_cell(ws_inp, r, 2, "Solar Only — RSC ($/yr)", black_font)
style_cell(ws_inp, r, 3, 1744, blue_font, dollar_fmt, yellow_fill)
style_cell(ws_inp, r, 4, 1113, blue_font, dollar_fmt, yellow_fill)
style_cell(ws_inp, r, 5, '=C30-D30', black_font, dollar_fmt)
r += 1
style_cell(ws_inp, r, 2, "Solar Only — Bridge ($/yr)", black_font)
style_cell(ws_inp, r, 3, 1624, blue_font, dollar_fmt, yellow_fill)
style_cell(ws_inp, r, 4, 1121, blue_font, dollar_fmt, yellow_fill)
style_cell(ws_inp, r, 5, '=C31-D31', black_font, dollar_fmt)

# Store key cell references for later use
# C5=PV+Storage cost, C6=Solar cost
# C12=interest, C13=discount, C14=term, C16=escalation
# C19=ITC rate, C20=ITC PV+S, C21=ITC Solar, C22=PowerPair, C23=EnergyWise
# C28=RSC no sys PV+S, D28=RSC w/ sys PV+S
# C29=Bridge no sys PV+S, D29=Bridge w/ sys PV+S
# C30=RSC no sys Solar, D30=RSC w/ sys Solar
# C31=Bridge no sys Solar, D31=Bridge w/ sys Solar

print("Inputs sheet created.")

# ================================================================
# HELPER: Build a Cash Flow Sheet
# ================================================================
def build_cashflow_sheet(ws, system_label, rate_label, 
                         cost_cell, bill_no_sys_cell, bill_w_sys_cell,
                         has_battery, scenarios):
    """
    Build a 25-year cash flow sheet with multiple scenarios.
    
    scenarios: list of dicts with keys:
      - name: scenario label
      - has_itc: bool
      - has_state: bool (PowerPair + EnergyWise)
    """
    ws.sheet_properties.tabColor = "70AD47" if has_battery else "ED7D31"
    
    # Column widths
    ws.column_dimensions['A'].width = 28
    for ci in range(2, 30):
        ws.column_dimensions[get_column_letter(ci)].width = 12
    
    # Title
    style_cell(ws, 1, 1, f"{system_label} — {rate_label} Rate", title_font)
    style_cell(ws, 2, 1, "All formulas reference Inputs sheet. Blue = input, Black = formula, Green = cross-sheet.", 
               Font(name='Arial', italic=True, size=9, color='666666'))
    
    current_row = 4
    
    for sc_idx, sc in enumerate(scenarios):
        start_row = current_row
        sc_name = sc['name']
        has_itc = sc['has_itc']
        has_state = sc['has_state']
        
        # Scenario header
        style_cell(ws, current_row, 1, f"Scenario {sc_idx+1}: {sc_name}", section_font)
        current_row += 1
        
        # Year row
        style_cell(ws, current_row, 1, "Year", header_font, None, light_blue_fill)
        for yr in range(26):
            cl = yr + 2  # B=0, C=1, ..., AA=25
            style_cell(ws, current_row, cl, yr, header_font, yr_fmt, light_blue_fill)
        # Total and NPV columns
        style_cell(ws, current_row, 28, "Total ($)", header_font, None, light_blue_fill)
        style_cell(ws, current_row, 29, "NPV ($)", header_font, None, light_blue_fill)
        yr_row = current_row
        current_row += 1
        
        # --- Debt section ---
        # Row: Initial Debt
        r_debt = current_row
        style_cell(ws, r_debt, 1, "Debt Balance ($)", black_font)
        # Year 0: system cost minus upfront incentives
        if has_itc and has_state and has_battery:
            # ITC + PowerPair reduce principal
            ws.cell(row=r_debt, column=2, value=f'=Inputs!{cost_cell}-Inputs!C20-Inputs!C22').font = green_font
        elif has_itc and has_state and not has_battery:
            ws.cell(row=r_debt, column=2, value=f'=Inputs!{cost_cell}-Inputs!C21').font = green_font
        elif not has_itc and has_state and has_battery:
            ws.cell(row=r_debt, column=2, value=f'=Inputs!{cost_cell}-Inputs!C22').font = green_font
        elif has_itc and not has_state:
            if has_battery:
                ws.cell(row=r_debt, column=2, value=f'=Inputs!{cost_cell}-Inputs!C20').font = green_font
            else:
                ws.cell(row=r_debt, column=2, value=f'=Inputs!{cost_cell}-Inputs!C21').font = green_font
        else:
            ws.cell(row=r_debt, column=2, value=f'=Inputs!{cost_cell}').font = green_font
        ws.cell(row=r_debt, column=2).number_format = dollar_fmt
        
        # Years 1-25: declining balance
        for yr in range(1, 26):
            cl = yr + 2
            prev = get_column_letter(cl - 1)
            ws.cell(row=r_debt, column=cl, 
                    value=f'=IF({prev}{r_debt}<=0,0,{prev}{r_debt}-{get_column_letter(cl)}{r_debt+2})').font = black_font
            ws.cell(row=r_debt, column=cl).number_format = dollar_fmt
        current_row += 1
        
        # Row: Interest Payment
        r_int = current_row
        style_cell(ws, r_int, 1, "Interest Payment ($)", black_font)
        ws.cell(row=r_int, column=2, value=0).font = black_font  # Year 0
        ws.cell(row=r_int, column=2).number_format = dollar_fmt
        for yr in range(1, 26):
            cl = yr + 2
            prev_debt = get_column_letter(cl - 1)
            ws.cell(row=r_int, column=cl, 
                    value=f'=IF({prev_debt}{r_debt}<=0,0,{prev_debt}{r_debt}*Inputs!$C$12)').font = black_font
            ws.cell(row=r_int, column=cl).number_format = dollar_fmt
        current_row += 1
        
        # Row: Principal Payment
        r_prin = current_row
        style_cell(ws, r_prin, 1, "Principal Payment ($)", black_font)
        ws.cell(row=r_prin, column=2, value=0).font = black_font
        ws.cell(row=r_prin, column=2).number_format = dollar_fmt
        for yr in range(1, 26):
            cl = yr + 2
            ws.cell(row=r_prin, column=cl,
                    value=f'=IF(B{r_debt}<=0,0,{get_column_letter(cl)}{r_int+1}-{get_column_letter(cl)}{r_int})').font = black_font
            ws.cell(row=r_prin, column=cl).number_format = dollar_fmt
        current_row += 1
        
        # Row: Total P&I
        r_pi = current_row
        style_cell(ws, r_pi, 1, "Total Payment (P&I) ($)", black_font)
        ws.cell(row=r_pi, column=2, value=0).font = black_font
        ws.cell(row=r_pi, column=2).number_format = dollar_fmt
        for yr in range(1, 26):
            cl = yr + 2
            ws.cell(row=r_pi, column=cl,
                    value=f'=IF(B{r_debt}<=0,0,ABS(PMT(Inputs!$C$12,Inputs!$C$14,B{r_debt},0)))').font = black_font
            ws.cell(row=r_pi, column=cl).number_format = dollar_fmt
        # Total and NPV
        ws.cell(row=r_pi, column=28, value=f'=SUM(B{r_pi}:AA{r_pi})').font = black_font
        ws.cell(row=r_pi, column=28).number_format = dollar_fmt
        ws.cell(row=r_pi, column=29, 
                value=f'=B{r_pi}+NPV(Inputs!$C$13,C{r_pi}:AA{r_pi})').font = black_font
        ws.cell(row=r_pi, column=29).number_format = dollar_fmt
        current_row += 1
        
        # --- Blank separator ---
        current_row += 1
        
        # Row: Elec Bill without system
        r_bill_no = current_row
        style_cell(ws, r_bill_no, 1, "Elec. Bill — No System ($)", black_font)
        ws.cell(row=r_bill_no, column=2, value=0).font = black_font
        ws.cell(row=r_bill_no, column=2).number_format = dollar_fmt
        # Year 1: from inputs
        ws.cell(row=r_bill_no, column=3, value=f'=Inputs!{bill_no_sys_cell}').font = green_font
        ws.cell(row=r_bill_no, column=3).number_format = dollar_fmt
        # Years 2-25: escalate
        for yr in range(2, 26):
            cl = yr + 2
            prev = get_column_letter(cl - 1)
            ws.cell(row=r_bill_no, column=cl,
                    value=f'={prev}{r_bill_no}*(1+Inputs!$C$16)').font = black_font
            ws.cell(row=r_bill_no, column=cl).number_format = dollar_fmt
        current_row += 1
        
        # Row: Elec Bill with system
        r_bill_w = current_row
        style_cell(ws, r_bill_w, 1, "Elec. Bill — With System ($)", black_font)
        ws.cell(row=r_bill_w, column=2, value=0).font = black_font
        ws.cell(row=r_bill_w, column=2).number_format = dollar_fmt
        ws.cell(row=r_bill_w, column=3, value=f'=Inputs!{bill_w_sys_cell}').font = green_font
        ws.cell(row=r_bill_w, column=3).number_format = dollar_fmt
        for yr in range(2, 26):
            cl = yr + 2
            prev = get_column_letter(cl - 1)
            ws.cell(row=r_bill_w, column=cl,
                    value=f'={prev}{r_bill_w}*(1+Inputs!$C$16)').font = black_font
            ws.cell(row=r_bill_w, column=cl).number_format = dollar_fmt
        current_row += 1
        
        # Row: Bill Savings
        r_sav = current_row
        style_cell(ws, r_sav, 1, "Bill Savings ($)", black_font)
        ws.cell(row=r_sav, column=2, value=0).font = black_font
        ws.cell(row=r_sav, column=2).number_format = dollar_fmt
        for yr in range(1, 26):
            cl = yr + 2
            ws.cell(row=r_sav, column=cl,
                    value=f'={get_column_letter(cl)}{r_bill_no}-{get_column_letter(cl)}{r_bill_w}').font = black_font
            ws.cell(row=r_sav, column=cl).number_format = dollar_fmt
        ws.cell(row=r_sav, column=28, value=f'=SUM(B{r_sav}:AA{r_sav})').font = black_font
        ws.cell(row=r_sav, column=28).number_format = dollar_fmt
        ws.cell(row=r_sav, column=29,
                value=f'=B{r_sav}+NPV(Inputs!$C$13,C{r_sav}:AA{r_sav})').font = black_font
        ws.cell(row=r_sav, column=29).number_format = dollar_fmt
        current_row += 1
        
        # Row: Incentive Savings (annual EnergyWise if applicable)
        r_inc = current_row
        style_cell(ws, r_inc, 1, "Incentive Savings ($)", black_font)
        ws.cell(row=r_inc, column=2, value=0).font = black_font
        ws.cell(row=r_inc, column=2).number_format = dollar_fmt
        for yr in range(1, 26):
            cl = yr + 2
            if has_state and has_battery:
                ws.cell(row=r_inc, column=cl, value=f'=Inputs!$C$23').font = green_font
            else:
                ws.cell(row=r_inc, column=cl, value=0).font = black_font
            ws.cell(row=r_inc, column=cl).number_format = dollar_fmt
        ws.cell(row=r_inc, column=28, value=f'=SUM(B{r_inc}:AA{r_inc})').font = black_font
        ws.cell(row=r_inc, column=28).number_format = dollar_fmt
        ws.cell(row=r_inc, column=29,
                value=f'=B{r_inc}+NPV(Inputs!$C$13,C{r_inc}:AA{r_inc})').font = black_font
        ws.cell(row=r_inc, column=29).number_format = dollar_fmt
        current_row += 1
        
        # Row: Total Savings
        r_tsav = current_row
        style_cell(ws, r_tsav, 1, "Total Savings ($)", header_font)
        for yr in range(0, 26):
            cl = yr + 2
            ws.cell(row=r_tsav, column=cl,
                    value=f'={get_column_letter(cl)}{r_sav}+{get_column_letter(cl)}{r_inc}').font = black_font
            ws.cell(row=r_tsav, column=cl).number_format = dollar_fmt
        ws.cell(row=r_tsav, column=28, value=f'=SUM(B{r_tsav}:AA{r_tsav})').font = black_font
        ws.cell(row=r_tsav, column=28).number_format = dollar_fmt
        ws.cell(row=r_tsav, column=29,
                value=f'=B{r_tsav}+NPV(Inputs!$C$13,C{r_tsav}:AA{r_tsav})').font = black_font
        ws.cell(row=r_tsav, column=29).number_format = dollar_fmt
        current_row += 1
        
        # Row: Net Savings (Total Savings - Total P&I)
        r_net = current_row
        style_cell(ws, r_net, 1, "Net Savings ($)", header_font)
        ws.cell(row=r_net, column=1).font = Font(name='Arial', bold=True, size=11, color='000000')
        for yr in range(0, 26):
            cl = yr + 2
            ws.cell(row=r_net, column=cl,
                    value=f'={get_column_letter(cl)}{r_tsav}-{get_column_letter(cl)}{r_pi}').font = black_font
            ws.cell(row=r_net, column=cl).number_format = dollar_fmt
        ws.cell(row=r_net, column=28, value=f'=SUM(B{r_net}:AA{r_net})').font = black_font
        ws.cell(row=r_net, column=28).number_format = dollar_fmt
        ws.cell(row=r_net, column=29,
                value=f'=B{r_net}+NPV(Inputs!$C$13,C{r_net}:AA{r_net})').font = black_font
        ws.cell(row=r_net, column=29).number_format = dollar_fmt
        
        # Bottom border for scenario
        for cl in range(1, 30):
            ws.cell(row=r_net, column=cl).border = bottom_border
        
        current_row += 2  # gap between scenarios
    
    return ws

# ================================================================
# BUILD CASH FLOW SHEETS
# ================================================================

# Scenario definitions
scenarios_pvs = [
    {"name": "100% Debt — FedITC + State Incentives", "has_itc": True, "has_state": True},
    {"name": "100% Debt — State Incentives Only (No FedITC)", "has_itc": False, "has_state": True},
    {"name": "100% Debt — No Incentives", "has_itc": False, "has_state": False},
]

scenarios_solar = [
    {"name": "100% Debt — With FedITC", "has_itc": True, "has_state": False},
    {"name": "100% Debt — No Incentives", "has_itc": False, "has_state": False},
]

# PV+Storage RSC
ws1 = wb.create_sheet("PV+Storage RSC")
build_cashflow_sheet(ws1, "PV + Storage", "RSC",
                     "C5", "C28", "D28", True, scenarios_pvs)
print("PV+Storage RSC sheet created.")

# PV+Storage Bridge
ws2 = wb.create_sheet("PV+Storage Bridge")
build_cashflow_sheet(ws2, "PV + Storage", "Bridge (TOU)",
                     "C5", "C29", "D29", True, scenarios_pvs)
print("PV+Storage Bridge sheet created.")

# Solar Only RSC
ws3 = wb.create_sheet("Solar Only RSC")
build_cashflow_sheet(ws3, "Solar Only", "RSC",
                     "C6", "C30", "D30", False, scenarios_solar)
print("Solar Only RSC sheet created.")

# Solar Only Bridge
ws4 = wb.create_sheet("Solar Only Bridge")
build_cashflow_sheet(ws4, "Solar Only", "Bridge (TOU)",
                     "C6", "C31", "D31", False, scenarios_solar)
print("Solar Only Bridge sheet created.")

# ================================================================
# SHEET 6: SUMMARY DASHBOARD
# ================================================================
ws_sum = wb.create_sheet("Summary")
ws_sum.sheet_properties.tabColor = "7030A0"

ws_sum.column_dimensions['A'].width = 14
ws_sum.column_dimensions['B'].width = 30
ws_sum.column_dimensions['C'].width = 18
ws_sum.column_dimensions['D'].width = 16
ws_sum.column_dimensions['E'].width = 16
ws_sum.column_dimensions['F'].width = 18
ws_sum.column_dimensions['G'].width = 18

style_cell(ws_sum, 1, 1, "NC Solar Financing — Summary Dashboard", title_font)
style_cell(ws_sum, 2, 1, "All values auto-update from cash flow sheets", Font(name='Arial', italic=True, size=9, color='666666'))

# Headers
r = 4
for ci, h in enumerate(["System", "Scenario", "Rate Structure", "System Cost ($)", "25yr NPV ($)", "Year 1 Net Savings ($)", "Monthly Net Savings Y1 ($)"], 1):
    style_cell(ws_sum, r, ci, h, header_font, None, light_blue_fill)

# Summary rows - reference the Net Savings NPV (column AC) from each scenario
# Each cash flow sheet has scenarios starting at different rows
# PV+Storage RSC: Scenario 1 net savings at row 17, Sc2 at row 33, Sc3 at row 49
# PV+Storage Bridge: same rows
# Solar Only: Sc1 net savings at row 17, Sc2 at row 33

# Net savings NPV is in column AC of the net savings row for each scenario
# Need to figure out exact rows. From the build function:
# Each scenario takes: 1 header + 1 year + debt(4) + blank(1) + bills(3) + inc(1) + total(1) + net(1) + gap(2) = 14 rows
# Sc1 starts at row 4: net savings = 4+1+1+4+1+3+1+1+0 = row 16... let me count:
# Row 4: scenario header
# Row 5: year row
# Row 6: debt balance
# Row 7: interest
# Row 8: principal
# Row 9: total P&I
# Row 10: blank
# Row 11: bill no sys
# Row 12: bill w sys
# Row 13: bill savings
# Row 14: incentive savings
# Row 15: total savings
# Row 16: net savings
# gap: 17-18
# Sc2: starts row 18
# Row 18: header, Row 19: year, ... Row 30: net savings
# Sc3: starts row 32
# ... Row 44: net savings (for 3 scenario sheets)
# For 2-scenario sheets: Sc1 net at 16, Sc2 net at 30

rows_3sc = [16, 30, 44]  # Net savings rows for 3-scenario sheets
rows_2sc = [16, 30]       # Net savings rows for 2-scenario sheets

summary_data = [
    # (System, Scenario, Rate, Cost ref, NPV ref, Y1 Net ref, Sheet)
    ("PV+Storage", "FedITC + State", "RSC", "='PV+Storage RSC'!B6", f"='PV+Storage RSC'!AC{rows_3sc[0]}", f"='PV+Storage RSC'!C{rows_3sc[0]}", "='PV+Storage RSC'!C{0}/12".format(rows_3sc[0])),
    ("PV+Storage", "State Only", "RSC", "='PV+Storage RSC'!B6", f"='PV+Storage RSC'!AC{rows_3sc[1]}", f"='PV+Storage RSC'!C{rows_3sc[1]}", "='PV+Storage RSC'!C{0}/12".format(rows_3sc[1])),
    ("PV+Storage", "No Incentives", "RSC", "='PV+Storage RSC'!B6", f"='PV+Storage RSC'!AC{rows_3sc[2]}", f"='PV+Storage RSC'!C{rows_3sc[2]}", "='PV+Storage RSC'!C{0}/12".format(rows_3sc[2])),
    ("PV+Storage", "FedITC + State", "Bridge", "='PV+Storage Bridge'!B6", f"='PV+Storage Bridge'!AC{rows_3sc[0]}", f"='PV+Storage Bridge'!C{rows_3sc[0]}", "='PV+Storage Bridge'!C{0}/12".format(rows_3sc[0])),
    ("PV+Storage", "State Only", "Bridge", "='PV+Storage Bridge'!B6", f"='PV+Storage Bridge'!AC{rows_3sc[1]}", f"='PV+Storage Bridge'!C{rows_3sc[1]}", "='PV+Storage Bridge'!C{0}/12".format(rows_3sc[1])),
    ("PV+Storage", "No Incentives", "Bridge", "='PV+Storage Bridge'!B6", f"='PV+Storage Bridge'!AC{rows_3sc[2]}", f"='PV+Storage Bridge'!C{rows_3sc[2]}", "='PV+Storage Bridge'!C{0}/12".format(rows_3sc[2])),
    ("Solar Only", "With FedITC", "RSC", "='Solar Only RSC'!B6", f"='Solar Only RSC'!AC{rows_2sc[0]}", f"='Solar Only RSC'!C{rows_2sc[0]}", "='Solar Only RSC'!C{0}/12".format(rows_2sc[0])),
    ("Solar Only", "No Incentives", "RSC", "='Solar Only RSC'!B6", f"='Solar Only RSC'!AC{rows_2sc[1]}", f"='Solar Only RSC'!C{rows_2sc[1]}", "='Solar Only RSC'!C{0}/12".format(rows_2sc[1])),
    ("Solar Only", "With FedITC", "Bridge", "='Solar Only Bridge'!B6", f"='Solar Only Bridge'!AC{rows_2sc[0]}", f"='Solar Only Bridge'!C{rows_2sc[0]}", "='Solar Only Bridge'!C{0}/12".format(rows_2sc[0])),
    ("Solar Only", "No Incentives", "Bridge", "='Solar Only Bridge'!B6", f"='Solar Only Bridge'!AC{rows_2sc[1]}", f"='Solar Only Bridge'!C{rows_2sc[1]}", "='Solar Only Bridge'!C{0}/12".format(rows_2sc[1])),
]

for i, (sys, sc, rate, cost_ref, npv_ref, y1_ref, mo_ref) in enumerate(summary_data):
    r = 5 + i
    style_cell(ws_sum, r, 1, sys, black_font)
    style_cell(ws_sum, r, 2, sc, black_font)
    style_cell(ws_sum, r, 3, rate, black_font)
    ws_sum.cell(row=r, column=4, value=cost_ref).font = green_font
    ws_sum.cell(row=r, column=4).number_format = dollar_fmt
    ws_sum.cell(row=r, column=5, value=npv_ref).font = green_font
    ws_sum.cell(row=r, column=5).number_format = dollar_fmt
    ws_sum.cell(row=r, column=6, value=y1_ref).font = green_font
    ws_sum.cell(row=r, column=6).number_format = dollar_fmt
    ws_sum.cell(row=r, column=7, value=mo_ref).font = green_font
    ws_sum.cell(row=r, column=7).number_format = dollar_fmt

print("Summary sheet created.")

# ================================================================
# SAVE
# ================================================================
output_path = HERE / "NC_Solar_Financing_Model.xlsx"
wb.save(output_path)
print(f"Saved: {output_path}")
