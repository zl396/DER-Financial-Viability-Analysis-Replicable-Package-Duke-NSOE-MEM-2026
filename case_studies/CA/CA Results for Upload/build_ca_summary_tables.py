"""
Build CA Summary Tables Excel — matching NC/MA format.
Two sheets: Guide + CA Results
"""

import json
from pathlib import Path
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

HERE = Path(__file__).resolve().parent
PYSAM_RESULTS = HERE.parent / "PySAM_outputs" / "CA_pysam_results.json"
with open(PYSAM_RESULTS) as f:
    results = json.load(f)
meta = results["_meta"]

def pmt(rate, nper, pv):
    return pv * rate * (1 + rate)**nper / ((1 + rate)**nper - 1)

def npv_calc(rate, cashflows):
    return sum(cf / (1 + rate)**t for t, cf in enumerate(cashflows))

r = 0.0724; n = 25; dr = 0.064; esc = 0.025

wb = Workbook()

# Styles
blk = Font(name='Arial', size=10)
bld = Font(name='Arial', bold=True, size=10)
ttl = Font(name='Arial', bold=True, size=14)
note = Font(name='Arial', italic=True, size=9, color='666666')
hdr_font = Font(name='Arial', bold=True, size=10, color='FFFFFF')
hdr_fill = PatternFill('solid', fgColor='4472C4')
green_font = Font(name='Arial', size=10, color='008000')
red_font = Font(name='Arial', size=10, color='CC0000')
lgr = PatternFill('solid', fgColor='F2F2F2')
dol = '$#,##0;($#,##0);"-"'

def sc(ws, r, c, v, font=blk, fmt=None, fill=None):
    cell = ws.cell(row=r, column=c, value=v)
    cell.font = font
    if fmt: cell.number_format = fmt
    if fill: cell.fill = fill
    return cell

# ================================================================
# GUIDE SHEET
# ================================================================
ws = wb.active
ws.title = "Guide"
ws.column_dimensions['A'].width = 18
ws.column_dimensions['B'].width = 90

sc(ws, 1, 1, "CA Summary Tables — Guide", ttl)

sc(ws, 3, 1, "Question answered:", bld)
sc(ws, 3, 2, "Under California's Net Billing Tariff (NBT), what is the 25-year financial return of residential solar and solar+storage across PG&E, SCE, and SDG&E territories?", blk)

sc(ws, 5, 1, "Column", bld); sc(ws, 5, 2, "Meaning", bld)
sc(ws, 6, 1, "Scenario", blk)
sc(ws, 6, 2, "System + utility combination. Solar Only = PV only. PV+Storage = PV + 13.5 kWh battery with TOU-aware dispatch (charges during super off-peak, discharges during peak).", blk)
sc(ws, 7, 1, "Utility", blk)
sc(ws, 7, 2, "PG&E (E-ELEC rate, Sacramento proxy), SCE (TOU-D-PRIME, Riverside proxy), SDG&E (EV-TOU-5, San Diego proxy). Each utility has different TOU rates and ACC export compensation.", blk)
sc(ws, 8, 1, "Bill w/o", blk)
sc(ws, 8, 2, "Annual electricity bill BEFORE installing DER. Same for Solar and PV+Storage within a utility.", blk)
sc(ws, 9, 1, "Bill Savings Y1", blk)
sc(ws, 9, 2, "Year 1 reduction in electricity bill. = Bill w/o - Bill w/ system. From PySAM with Utilityrate5 (TOU import rates + hourly ACC export rates).", blk)
sc(ws, 10, 1, "PMT", blk)
sc(ws, 10, 2, "Annual loan payment = PMT(7.24%, 25yr, system cost). No ITC or state incentives applied (all expired/closed).", blk)
sc(ws, 11, 1, "Monthly Net", blk)
sc(ws, 11, 2, "(Bill Savings - PMT) / 12. Negative = customer pays out of pocket. Positive = net monthly income.", blk)
sc(ws, 12, 1, "25yr NPV", blk)
sc(ws, 12, 2, "Net Present Value over 25 years at 6.4% discount rate. Bill savings escalate at 2.5%/yr. 100% debt financed. Positive = project creates value.", blk)
sc(ws, 13, 1, "Simple Payback", blk)
sc(ws, 13, 2, "System cost / Y1 bill savings. Does not account for financing cost or escalation.", blk)

sc(ws, 15, 1, "Key Parameters:", bld)
sc(ws, 16, 1, "  \u2022", blk); sc(ws, 16, 2, f"PV: {meta['pv_kw']} kW, ${meta['solar_cost_per_kw']:,}/kW = ${meta['solar_cost']:,.0f} (Tao Sun CA, 2025)", blk)
sc(ws, 17, 1, "  \u2022", blk); sc(ws, 17, 2, f"Battery: {meta['battery_kwh']} kWh, ${meta['batt_cost_per_kwh']:,}/kWh = ${meta['batt_cost']:,.0f} (Tao Sun CA, 2025)", blk)
sc(ws, 18, 1, "  \u2022", blk); sc(ws, 18, 2, "Loan: 7.24%, 25yr, 100% debt financed", blk)
sc(ws, 19, 1, "  \u2022", blk); sc(ws, 19, 2, "Federal ITC: 0% (expired Dec 2025)", blk)
sc(ws, 20, 1, "  \u2022", blk); sc(ws, 20, 2, "SGIP/RSSE: $0 (SGIP closed Dec 2025, RSSE fully reserved)", blk)
sc(ws, 21, 1, "  \u2022", blk); sc(ws, 21, 2, "Export compensation: Avoided Cost Calculator (ACC) hourly rates, 2026 vintage, ~$0.09/kWh annual average", blk)
sc(ws, 22, 1, "  \u2022", blk); sc(ws, 22, 2, "Load: ResStock county-proxy profiles (Sacramento 8,917 kWh, Riverside 8,431 kWh, San Diego 7,220 kWh)", blk)

print("Guide sheet done.")

# ================================================================
# CA RESULTS SHEET
# ================================================================
ws2 = wb.create_sheet("CA Results")
ws2.column_dimensions['A'].width = 22
ws2.column_dimensions['B'].width = 14
ws2.column_dimensions['C'].width = 12
ws2.column_dimensions['D'].width = 16
ws2.column_dimensions['E'].width = 10
ws2.column_dimensions['F'].width = 14
ws2.column_dimensions['G'].width = 12
ws2.column_dimensions['H'].width = 14

sc(ws2, 1, 1, "CA DER Financial Results — No Incentives (Phase A)", ttl)
sc(ws2, 2, 1, f"System: {meta['pv_kw']} kW PV (${meta['solar_cost']:,.0f}), {meta['battery_kwh']} kWh battery (${meta['batt_cost']:,.0f}). Loan: 7.24%, 25yr. Export: NBT ACC hourly rates.", note)

# Headers
headers = ["Scenario", "Utility", "Bill w/o", "Bill Savings Y1", "PMT", "Monthly Net", "25yr NPV", "Simple Payback"]
for i, h in enumerate(headers):
    sc(ws2, 4, i+1, h, hdr_font, fill=hdr_fill)

# Build rows
scenarios_order = [
    ("Solar Only", "PGE", "PG&E", meta["solar_cost"]),
    ("Solar Only", "SCE", "SCE", meta["solar_cost"]),
    ("Solar Only", "SDGE", "SDG&E", meta["solar_cost"]),
    ("PV+Storage", "PGE", "PG&E", meta["total_pvs_cost"]),
    ("PV+Storage", "SCE", "SCE", meta["total_pvs_cost"]),
    ("PV+Storage", "SDGE", "SDG&E", meta["total_pvs_cost"]),
]

row = 5
for sys_label, util_key, util_name, cost in scenarios_order:
    key = f"{util_key}_{'SolarOnly' if 'Solar Only' in sys_label else 'SolarBattery'}"
    d = results[key]
    yr1_sav = d["bill_wo"] - d["bill_w"]
    annual_pmt = pmt(r, n, cost)
    yr1_net = yr1_sav - annual_pmt
    mo_net = yr1_net / 12

    flows = [yr1_sav * (1+esc)**(y-1) - annual_pmt for y in range(1,26)]
    nv = npv_calc(dr, [0] + flows)

    simple_pb = cost / yr1_sav if yr1_sav > 0 else 999
    pb_str = f"{simple_pb:.1f} yr"

    # Color code monthly net
    mo_font = green_font if mo_net >= 0 else red_font
    npv_font = green_font if nv >= 0 else red_font

    fill = lgr if row % 2 == 0 else None

    sc(ws2, row, 1, sys_label, blk, fill=fill)
    sc(ws2, row, 2, util_name, blk, fill=fill)
    sc(ws2, row, 3, round(d["bill_wo"]), blk, dol, fill)
    sc(ws2, row, 4, round(yr1_sav), blk, dol, fill)
    sc(ws2, row, 5, round(annual_pmt), blk, dol, fill)
    sc(ws2, row, 6, round(mo_net), mo_font, dol, fill)
    sc(ws2, row, 7, round(nv), npv_font, dol, fill)
    sc(ws2, row, 8, pb_str, blk, fill=fill)
    row += 1

# Takeaway
row += 1
sc(ws2, row, 1, "Takeaway:", bld)
sc(ws2, row, 2, "Under NBT with no incentives, solar-only is NPV-positive for PG&E (+$2,649) and SDG&E (+$1,839) but marginal for SCE (-$39). PV+Storage is only viable at SDG&E (+$3,064) where the extreme TOU spread ($0.12-$0.80/kWh) enables $1,100/yr in battery arbitrage value. At PG&E and SCE, battery financing cost ($1,291/yr) exceeds incremental savings ($614-$737/yr), making storage uneconomic without ITC or SGIP.", blk)

# ================================================================
# BATTERY VALUE SHEET
# ================================================================
ws3 = wb.create_sheet("Battery Value")
ws3.column_dimensions['A'].width = 14
ws3.column_dimensions['B'].width = 16
ws3.column_dimensions['C'].width = 20
ws3.column_dimensions['D'].width = 18
ws3.column_dimensions['E'].width = 14
ws3.column_dimensions['F'].width = 16

sc(ws3, 1, 1, "Battery Incremental Value Analysis", ttl)
sc(ws3, 2, 1, f"Battery: {meta['battery_kwh']} kWh / {meta['battery_kw']} kW, ${meta['batt_cost']:,.0f}. Incremental PMT = ${pmt(r,n,meta['batt_cost']):,.0f}/yr.", note)

headers3 = ["Utility", "Rate Schedule", "Incremental Savings", "Incremental PMT", "Y1 Net", "25yr \u0394 NPV"]
for i, h in enumerate(headers3):
    sc(ws3, 4, i+1, h, hdr_font, fill=hdr_fill)

batt_pmt_annual = pmt(r, n, meta["batt_cost"])

for i, (util_key, util_name, rate_name) in enumerate([
    ("PGE", "PG&E", "E-ELEC"),
    ("SCE", "SCE", "TOU-D-PRIME"),
    ("SDGE", "SDG&E", "EV-TOU-5"),
]):
    solar = results[f"{util_key}_SolarOnly"]
    batt = results[f"{util_key}_SolarBattery"]
    incr_sav = (batt["bill_wo"] - batt["bill_w"]) - (solar["bill_wo"] - solar["bill_w"])
    incr_net = incr_sav - batt_pmt_annual

    solar_sav = solar["bill_wo"] - solar["bill_w"]
    pvs_sav = batt["bill_wo"] - batt["bill_w"]
    solar_flows = [solar_sav*(1+esc)**(y-1) - pmt(r,n,meta["solar_cost"]) for y in range(1,26)]
    pvs_flows = [pvs_sav*(1+esc)**(y-1) - pmt(r,n,meta["total_pvs_cost"]) for y in range(1,26)]
    delta_npv = npv_calc(dr, [0]+pvs_flows) - npv_calc(dr, [0]+solar_flows)

    row = 5 + i
    fill = lgr if row % 2 == 0 else None
    npv_font_use = green_font if delta_npv >= 0 else red_font
    net_font = green_font if incr_net >= 0 else red_font

    sc(ws3, row, 1, util_name, blk, fill=fill)
    sc(ws3, row, 2, rate_name, blk, fill=fill)
    sc(ws3, row, 3, f"${incr_sav:,.0f}/yr", blk, fill=fill)
    sc(ws3, row, 4, f"${batt_pmt_annual:,.0f}/yr", blk, fill=fill)
    sc(ws3, row, 5, round(incr_net), net_font, dol, fill)
    sc(ws3, row, 6, round(delta_npv), npv_font_use, dol, fill)

row = 9
sc(ws3, row, 1, "Takeaway:", bld)
sc(ws3, row, 2, "SDG&E's EV-TOU-5 is the only rate where battery storage breaks even without incentives, driven by the nation's widest residential TOU spread ($0.68/kWh summer peak-to-trough). At PG&E and SCE, the spread ($0.22 and $0.33 respectively) generates insufficient arbitrage to cover battery financing at current costs and interest rates.", blk)

# ================================================================
# RATE STRUCTURE SHEET
# ================================================================
ws4 = wb.create_sheet("Rate Structures")
ws4.column_dimensions['A'].width = 14
ws4.column_dimensions['B'].width = 16
ws4.column_dimensions['C'].width = 16
ws4.column_dimensions['D'].width = 18
ws4.column_dimensions['E'].width = 22
ws4.column_dimensions['F'].width = 14
ws4.column_dimensions['G'].width = 18
ws4.column_dimensions['H'].width = 22
ws4.column_dimensions['I'].width = 14
ws4.column_dimensions['J'].width = 14

sc(ws4, 1, 1, "CA NBT Mandatory Rate Structures (2026)", ttl)
sc(ws4, 2, 1, "All NBT solar customers must enroll in these electrification TOU rates. Export compensated at hourly ACC values.", note)

headers4 = ["Utility", "Rate Schedule", "Summer Peak", "Summer Off-Peak", "Summer Super Off-Peak",
            "Winter Peak", "Winter Off-Peak", "Winter Super Off-Peak", "Fixed/mo", "Avg ACC Export"]
for i, h in enumerate(headers4):
    sc(ws4, 4, i+1, h, hdr_font, fill=hdr_fill)

rate_data = [
    ("PG&E", "E-ELEC", "$0.552", "$0.334", "—", "$0.321", "$0.285", "—", "$24", "$0.097"),
    ("SCE", "TOU-D-PRIME", "$0.590", "$0.260", "—", "$0.560", "$0.240", "$0.240", "$24", "$0.092"),
    ("SDG&E", "EV-TOU-5", "$0.800", "$0.502", "$0.124", "$0.529", "$0.473", "$0.117", "$24", "$0.090"),
]

for i, data in enumerate(rate_data):
    row = 5 + i
    fill = lgr if row % 2 == 0 else None
    for j, val in enumerate(data):
        sc(ws4, row, j+1, val, blk, fill=fill)

row = 9
sc(ws4, row, 1, "Takeaway:", bld)
sc(ws4, row, 2, "SDG&E's EV-TOU-5 has the widest peak-to-trough spread in the nation ($0.80 vs $0.12 = $0.68/kWh), making it the strongest market signal for battery storage. PG&E's E-ELEC has a narrow spread ($0.22 summer, $0.04 winter), limiting TOU arbitrage value. All three utilities share ~$0.09/kWh average ACC export rates, representing a 70-85% reduction from NEM 2.0 retail-rate exports.", blk)

# ================================================================
# SAVE
# ================================================================
out = HERE / "CA_Summary_Tables.xlsx"
wb.save(out)
print(f"Saved: {out}")
print(f"Sheets: {wb.sheetnames}")
