"""
Build CA Solar Financing Excel Model — 3 Utilities × 2 Systems = 6 scenarios.

Aligned with NC/MA pattern:
  Inputs sheet → Cash Flow sheets (25-year) → Summary dashboard
  No incentives (Phase A). SGIP sensitivity to be added later (Phase C).
"""

import json
from pathlib import Path
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Load PySAM results
HERE = Path(__file__).resolve().parent
PYSAM_RESULTS = HERE.parent / "PySAM_outputs" / "CA_pysam_results.json"
with open(PYSAM_RESULTS) as f:
    results = json.load(f)
meta = results["_meta"]

wb = Workbook()

# === STYLES ===
blue = Font(name='Arial', color='0000FF', size=10)
blk = Font(name='Arial', color='000000', size=10)
grn = Font(name='Arial', color='008000', size=10)
hdr = Font(name='Arial', bold=True, size=10, color='000000')
sec = Font(name='Arial', bold=True, size=12, color='000000')
ttl = Font(name='Arial', bold=True, size=14, color='000000')
note = Font(name='Arial', italic=True, size=9, color='666666')
yel = PatternFill('solid', fgColor='FFFF00')
lgr = PatternFill('solid', fgColor='F2F2F2')
lbl = PatternFill('solid', fgColor='DCE6F1')
btm = Border(bottom=Side(style='medium'))
dol = '$#,##0;($#,##0);"-"'
dol2 = '$#,##0.00;($#,##0.00);"-"'
pct = '0.0%'
yrfmt = '0'


def sc(ws, r, c, v, f=blk, fmt=None, fill=None):
    cell = ws.cell(row=r, column=c, value=v)
    cell.font = f
    if fmt:
        cell.number_format = fmt
    if fill:
        cell.fill = fill
    return cell


# ================================================================
# INPUTS SHEET
# ================================================================
ws = wb.active
ws.title = "Inputs"
ws.sheet_properties.tabColor = "4472C4"
ws.column_dimensions['A'].width = 4
ws.column_dimensions['B'].width = 42
ws.column_dimensions['C'].width = 16
ws.column_dimensions['D'].width = 16
ws.column_dimensions['E'].width = 50

sc(ws, 1, 2, "CA Solar Financing Model — Input Parameters", ttl)
sc(ws, 2, 2, "Blue = editable  |  Black = formula  |  Green = cross-sheet link", note)

# ROW MAP:
# C5  = PV+Storage System Cost
# C6  = Solar Only System Cost
# C7  = Solar PV kW
# C8  = Battery kWh
# C11 = Interest Rate
# C12 = Discount Rate
# C13 = Loan Term
# C14 = Escalation
# C17 = ITC Rate (0% — expired)
# C18 = ITC PV+S
# C19 = ITC Solar
# C20 = SGIP/State Incentive (0 for Phase A)
# C25-C30 = Year 1 bills (without/with) for 6 scenarios

sc(ws, 4, 2, "SYSTEM PARAMETERS", sec)
sc(ws, 5, 2, "PV + Storage System Cost", blk)
sc(ws, 5, 3, meta["total_pvs_cost"], blue, dol, yel); sc(ws, 5, 4, "$", blk)
sc(ws, 5, 5, f"Solar ${meta['solar_cost']:,.0f} + Battery ${meta['batt_cost']:,.0f}", note)

sc(ws, 6, 2, "Solar Only System Cost", blk)
sc(ws, 6, 3, meta["solar_cost"], blue, dol, yel); sc(ws, 6, 4, "$", blk)
sc(ws, 6, 5, f"Source: Tao Sun et al. (2025), CA: ${meta['solar_cost_per_kw']}/kW", note)

sc(ws, 7, 2, "Solar PV Capacity", blk)
sc(ws, 7, 3, meta["pv_kw"], blue, '0.00', yel); sc(ws, 7, 4, "kW", blk)

sc(ws, 8, 2, "Battery Storage Capacity", blk)
sc(ws, 8, 3, meta["battery_kwh"], blue, '0.0', yel); sc(ws, 8, 4, "kWh", blk)

# Financial
sc(ws, 10, 2, "FINANCIAL PARAMETERS", sec)
sc(ws, 11, 2, "Debt Interest Rate", blk)
sc(ws, 11, 3, meta["loan_rate"], blue, pct, yel)

sc(ws, 12, 2, "Customer Discount Rate", blk)
sc(ws, 12, 3, meta["discount_rate"], blue, pct, yel)

sc(ws, 13, 2, "Loan Term", blk)
sc(ws, 13, 3, meta["loan_term_years"], blue, yrfmt, yel); sc(ws, 13, 4, "years", blk)

sc(ws, 14, 2, "Electricity Cost Escalation Rate", blk)
sc(ws, 14, 3, meta["escalation_rate"], blue, pct, yel)

# Incentives
sc(ws, 16, 2, "INCENTIVES", sec)
sc(ws, 17, 2, "Federal ITC Rate", blk)
sc(ws, 17, 3, 0.0, blue, pct, yel)
sc(ws, 17, 5, "Expired Dec 2025; set to 0% for current scenario", note)

sc(ws, 18, 2, "FedITC Amount — PV+Storage", blk)
sc(ws, 18, 3, '=C5*C17', blk, dol)

sc(ws, 19, 2, "FedITC Amount — Solar Only", blk)
sc(ws, 19, 3, '=C6*C17', blk, dol)

sc(ws, 20, 2, "State Battery Incentive (SGIP)", blk)
sc(ws, 20, 3, 0, blue, dol, yel)
sc(ws, 20, 5, "SGIP closed Dec 2025; $0 for Phase A, sensitivity in Phase C", note)

# Year 1 bills
sc(ws, 22, 2, "YEAR 1 ELECTRICITY BILLS (from PySAM)", sec)
sc(ws, 23, 3, "Without System", hdr, None, lgr)
sc(ws, 23, 4, "With System", hdr, None, lgr)
sc(ws, 23, 5, "Annual Savings", hdr, None, lgr)

# Row 25: PGE Solar      Row 26: PGE PV+S
# Row 27: SCE Solar       Row 28: SCE PV+S
# Row 29: SDGE Solar      Row 30: SDGE PV+S
bill_rows = {
    "PGE_SolarOnly": 25,
    "PGE_SolarBattery": 26,
    "SCE_SolarOnly": 27,
    "SCE_SolarBattery": 28,
    "SDGE_SolarOnly": 29,
    "SDGE_SolarBattery": 30,
}

labels = {
    "PGE_SolarOnly": "PG&E — Solar Only (E-ELEC)",
    "PGE_SolarBattery": "PG&E — PV+Storage (E-ELEC)",
    "SCE_SolarOnly": "SCE — Solar Only (TOU-D-PRIME)",
    "SCE_SolarBattery": "SCE — PV+Storage (TOU-D-PRIME)",
    "SDGE_SolarOnly": "SDG&E — Solar Only (EV-TOU-5)",
    "SDGE_SolarBattery": "SDG&E — PV+Storage (EV-TOU-5)",
}

for key, row in bill_rows.items():
    r = results[key]
    sc(ws, row, 2, labels[key], blk)
    sc(ws, row, 3, r["bill_wo"], blue, dol, yel)
    sc(ws, row, 4, r["bill_w"], blue, dol, yel)
    sc(ws, row, 5, f'=C{row}-D{row}', blk, dol)

# PV production metadata
sc(ws, 32, 2, "PV PRODUCTION (from PySAM)", sec)
sc(ws, 33, 2, "PG&E (Sacramento) Annual PV kWh", blk)
sc(ws, 33, 3, results["PGE_SolarOnly"]["pv_annual_kwh"], blk, '#,##0')
sc(ws, 34, 2, "SCE (Riverside) Annual PV kWh", blk)
sc(ws, 34, 3, results["SCE_SolarOnly"]["pv_annual_kwh"], blk, '#,##0')
sc(ws, 35, 2, "SDG&E (San Diego) Annual PV kWh", blk)
sc(ws, 35, 3, results["SDGE_SolarOnly"]["pv_annual_kwh"], blk, '#,##0')
sc(ws, 36, 2, "PG&E Load Annual kWh", blk)
sc(ws, 36, 3, results["PGE_SolarOnly"]["load_annual_kwh"], blk, '#,##0')
sc(ws, 37, 2, "SCE Load Annual kWh", blk)
sc(ws, 37, 3, results["SCE_SolarOnly"]["load_annual_kwh"], blk, '#,##0')
sc(ws, 38, 2, "SDG&E Load Annual kWh", blk)
sc(ws, 38, 3, results["SDGE_SolarOnly"]["load_annual_kwh"], blk, '#,##0')

print("Inputs sheet done.")


# ================================================================
# CASH FLOW BUILDER
# ================================================================
def build_cf(ws, title_str, cost_cell, bill_no_cell, bill_w_cell, has_batt, scenarios):
    """Build cash flow sheet with 25-year projection.

    scenarios: list of dicts with keys: name, has_itc, has_sgip
    """
    ws.sheet_properties.tabColor = "70AD47" if has_batt else "ED7D31"
    ws.column_dimensions['A'].width = 30
    for ci in range(2, 30):
        ws.column_dimensions[get_column_letter(ci)].width = 12

    sc(ws, 1, 1, title_str, ttl)
    sc(ws, 2, 1, "All formulas reference Inputs sheet", note)

    row = 4
    net_rows = []

    for si, s in enumerate(scenarios):
        sc(ws, row, 1, f"Scenario {si+1}: {s['name']}", sec)
        row += 1

        # Year header
        yr_row = row
        sc(ws, yr_row, 1, "Year", hdr, None, lbl)
        for y in range(26):
            sc(ws, yr_row, y + 2, y, hdr, yrfmt, lbl)
        sc(ws, yr_row, 28, "25yr Total", hdr, None, lbl)
        sc(ws, yr_row, 29, "25yr NPV", hdr, None, lbl)
        row += 1

        # Debt Balance
        r_db = row
        sc(ws, r_db, 1, "Debt Balance ($)", blk)
        parts = [f"Inputs!{cost_cell}"]
        if s["has_itc"]:
            itc_cell = "C18" if has_batt else "C19"
            parts.append(f"Inputs!{itc_cell}")
        if s.get("has_sgip") and has_batt:
            parts.append("Inputs!C20")
        yr0_formula = "=" + "-".join(parts)
        ws.cell(row=r_db, column=2, value=yr0_formula).font = grn
        ws.cell(row=r_db, column=2).number_format = dol
        for y in range(1, 26):
            c = y + 2
            p = get_column_letter(c - 1)
            ws.cell(row=r_db, column=c, value=f'=MAX(0,{p}{r_db}-{get_column_letter(c)}{r_db+2})').font = blk
            ws.cell(row=r_db, column=c).number_format = dol
        row += 1

        # Interest
        r_int = row
        sc(ws, r_int, 1, "Interest Payment ($)", blk)
        sc(ws, r_int, 2, 0, blk, dol)
        for y in range(1, 26):
            c = y + 2
            p = get_column_letter(c - 1)
            ws.cell(row=r_int, column=c, value=f'=IF({p}{r_db}<=0,0,{p}{r_db}*Inputs!$C$11)').font = blk
            ws.cell(row=r_int, column=c).number_format = dol
        row += 1

        # Principal
        r_pr = row
        sc(ws, r_pr, 1, "Principal Payment ($)", blk)
        sc(ws, r_pr, 2, 0, blk, dol)
        for y in range(1, 26):
            c = y + 2
            cl = get_column_letter(c)
            ws.cell(row=r_pr, column=c, value=f'=IF(B{r_db}<=0,0,{cl}{r_pr+1}-{cl}{r_int})').font = blk
            ws.cell(row=r_pr, column=c).number_format = dol
        row += 1

        # Total P&I
        r_pi = row
        sc(ws, r_pi, 1, "Total Payment — P&I ($)", hdr)
        sc(ws, r_pi, 2, 0, blk, dol)
        for y in range(1, 26):
            c = y + 2
            p = get_column_letter(c - 1)
            ws.cell(row=r_pi, column=c, value=f'=IF({p}{r_db}<=0,0,ABS(PMT(Inputs!$C$11,Inputs!$C$13,B{r_db},0)))').font = blk
            ws.cell(row=r_pi, column=c).number_format = dol
        ws.cell(row=r_pi, column=28, value=f'=SUM(C{r_pi}:AA{r_pi})').font = blk
        ws.cell(row=r_pi, column=28).number_format = dol
        ws.cell(row=r_pi, column=29, value=f'=NPV(Inputs!$C$12,C{r_pi}:AA{r_pi})').font = blk
        ws.cell(row=r_pi, column=29).number_format = dol
        row += 2

        # Bill without system
        r_bn = row
        sc(ws, r_bn, 1, "Elec. Bill — No System ($)", blk)
        sc(ws, r_bn, 2, 0, blk, dol)
        ws.cell(row=r_bn, column=3, value=f'=Inputs!{bill_no_cell}').font = grn
        ws.cell(row=r_bn, column=3).number_format = dol
        for y in range(2, 26):
            c = y + 2
            p = get_column_letter(c - 1)
            ws.cell(row=r_bn, column=c, value=f'={p}{r_bn}*(1+Inputs!$C$14)').font = blk
            ws.cell(row=r_bn, column=c).number_format = dol
        row += 1

        # Bill with system
        r_bw = row
        sc(ws, r_bw, 1, "Elec. Bill — With System ($)", blk)
        sc(ws, r_bw, 2, 0, blk, dol)
        ws.cell(row=r_bw, column=3, value=f'=Inputs!{bill_w_cell}').font = grn
        ws.cell(row=r_bw, column=3).number_format = dol
        for y in range(2, 26):
            c = y + 2
            p = get_column_letter(c - 1)
            ws.cell(row=r_bw, column=c, value=f'={p}{r_bw}*(1+Inputs!$C$14)').font = blk
            ws.cell(row=r_bw, column=c).number_format = dol
        row += 1

        # Bill Savings
        r_bs = row
        sc(ws, r_bs, 1, "Bill Savings ($)", blk)
        sc(ws, r_bs, 2, 0, blk, dol)
        for y in range(1, 26):
            c = y + 2
            cl = get_column_letter(c)
            ws.cell(row=r_bs, column=c, value=f'={cl}{r_bn}-{cl}{r_bw}').font = blk
            ws.cell(row=r_bs, column=c).number_format = dol
        ws.cell(row=r_bs, column=28, value=f'=SUM(C{r_bs}:AA{r_bs})').font = blk
        ws.cell(row=r_bs, column=28).number_format = dol
        ws.cell(row=r_bs, column=29, value=f'=NPV(Inputs!$C$12,C{r_bs}:AA{r_bs})').font = blk
        ws.cell(row=r_bs, column=29).number_format = dol
        row += 1

        # Incentive Savings (annual — placeholder for SGIP sensitivity)
        r_is = row
        sc(ws, r_is, 1, "Incentive Savings — Annual ($)", blk)
        for y in range(0, 26):
            c = y + 2
            ws.cell(row=r_is, column=c, value=0).font = blk
            ws.cell(row=r_is, column=c).number_format = dol
        ws.cell(row=r_is, column=28, value=f'=SUM(C{r_is}:AA{r_is})').font = blk
        ws.cell(row=r_is, column=28).number_format = dol
        ws.cell(row=r_is, column=29, value=f'=NPV(Inputs!$C$12,C{r_is}:AA{r_is})').font = blk
        ws.cell(row=r_is, column=29).number_format = dol
        row += 1

        # Total Savings
        r_ts = row
        sc(ws, r_ts, 1, "Total Savings ($)", hdr)
        for y in range(0, 26):
            c = y + 2
            cl = get_column_letter(c)
            ws.cell(row=r_ts, column=c, value=f'={cl}{r_bs}+{cl}{r_is}').font = blk
            ws.cell(row=r_ts, column=c).number_format = dol
        ws.cell(row=r_ts, column=28, value=f'=SUM(C{r_ts}:AA{r_ts})').font = blk
        ws.cell(row=r_ts, column=28).number_format = dol
        ws.cell(row=r_ts, column=29, value=f'=NPV(Inputs!$C$12,C{r_ts}:AA{r_ts})').font = blk
        ws.cell(row=r_ts, column=29).number_format = dol
        row += 1

        # Net Savings
        r_ns = row
        sc(ws, r_ns, 1, "Net Savings ($)", Font(name='Arial', bold=True, size=11, color='000000'))
        for y in range(0, 26):
            c = y + 2
            cl = get_column_letter(c)
            ws.cell(row=r_ns, column=c, value=f'={cl}{r_ts}-{cl}{r_pi}').font = blk
            ws.cell(row=r_ns, column=c).number_format = dol
        ws.cell(row=r_ns, column=28, value=f'=SUM(C{r_ns}:AA{r_ns})').font = blk
        ws.cell(row=r_ns, column=28).number_format = dol
        ws.cell(row=r_ns, column=29, value=f'=NPV(Inputs!$C$12,C{r_ns}:AA{r_ns})').font = blk
        ws.cell(row=r_ns, column=29).number_format = dol
        net_rows.append(r_ns)

        for c in range(1, 30):
            ws.cell(row=r_ns, column=c).border = btm
        row += 3

    return net_rows


# ================================================================
# BUILD SHEETS — One per utility, Solar and PV+S scenarios together
# ================================================================

# Scenario definitions (Phase A: No incentives only)
sc_no_incentives = [
    {"name": "No Incentives (ITC=0%, SGIP=$0)", "has_itc": False, "has_sgip": False},
]

# PG&E
ws_pge_s = wb.create_sheet("PGE Solar")
nr_pge_s = build_cf(ws_pge_s, "PG&E — Solar Only (E-ELEC)", "C6", "C25", "D25", False, sc_no_incentives)

ws_pge_b = wb.create_sheet("PGE PV+Storage")
nr_pge_b = build_cf(ws_pge_b, "PG&E — PV+Storage (E-ELEC)", "C5", "C26", "D26", True, sc_no_incentives)

# SCE
ws_sce_s = wb.create_sheet("SCE Solar")
nr_sce_s = build_cf(ws_sce_s, "SCE — Solar Only (TOU-D-PRIME)", "C6", "C27", "D27", False, sc_no_incentives)

ws_sce_b = wb.create_sheet("SCE PV+Storage")
nr_sce_b = build_cf(ws_sce_b, "SCE — PV+Storage (TOU-D-PRIME)", "C5", "C28", "D28", True, sc_no_incentives)

# SDG&E
ws_sdge_s = wb.create_sheet("SDGE Solar")
nr_sdge_s = build_cf(ws_sdge_s, "SDG&E — Solar Only (EV-TOU-5)", "C6", "C29", "D29", False, sc_no_incentives)

ws_sdge_b = wb.create_sheet("SDGE PV+Storage")
nr_sdge_b = build_cf(ws_sdge_b, "SDG&E — PV+Storage (EV-TOU-5)", "C5", "C30", "D30", True, sc_no_incentives)

print("All cash flow sheets done.")

# ================================================================
# SUMMARY DASHBOARD
# ================================================================
ws_s = wb.create_sheet("Summary")
ws_s.sheet_properties.tabColor = "7030A0"
ws_s.column_dimensions['A'].width = 12
ws_s.column_dimensions['B'].width = 20
ws_s.column_dimensions['C'].width = 18
ws_s.column_dimensions['D'].width = 16
ws_s.column_dimensions['E'].width = 14
ws_s.column_dimensions['F'].width = 18
ws_s.column_dimensions['G'].width = 18
ws_s.column_dimensions['H'].width = 18

sc(ws_s, 1, 1, "CA Solar Financing — Summary Dashboard (Phase A: No Incentives)", ttl)
sc(ws_s, 2, 1, "Green = cross-sheet formula  |  All scenarios: ITC=0%, SGIP=$0", note)

headers = ["Utility", "System", "Rate Schedule", "System Cost ($)",
           "Y1 Savings ($)", "25yr NPV ($)", "Y1 Net Monthly ($)"]
for i, h in enumerate(headers):
    sc(ws_s, 4, i + 1, h, hdr, None, lbl)

sheet_data = [
    ("PG&E", "Solar Only", "E-ELEC", "'PGE Solar'", nr_pge_s[0], "C6"),
    ("PG&E", "PV+Storage", "E-ELEC", "'PGE PV+Storage'", nr_pge_b[0], "C5"),
    ("SCE", "Solar Only", "TOU-D-PRIME", "'SCE Solar'", nr_sce_s[0], "C6"),
    ("SCE", "PV+Storage", "TOU-D-PRIME", "'SCE PV+Storage'", nr_sce_b[0], "C5"),
    ("SDG&E", "Solar Only", "EV-TOU-5", "'SDGE Solar'", nr_sdge_s[0], "C6"),
    ("SDG&E", "PV+Storage", "EV-TOU-5", "'SDGE PV+Storage'", nr_sdge_b[0], "C5"),
]

for i, (util, sys, rate, sheet, net_r, cost_c) in enumerate(sheet_data):
    r = 5 + i
    sc(ws_s, r, 1, util, blk)
    sc(ws_s, r, 2, sys, blk)
    sc(ws_s, r, 3, rate, blk)
    # System cost
    ws_s.cell(row=r, column=4, value=f'=Inputs!{cost_c}').font = grn
    ws_s.cell(row=r, column=4).number_format = dol
    # Y1 savings = Y1 net savings (total savings - P&I)
    ws_s.cell(row=r, column=5, value=f'={sheet}!C{net_r}').font = grn
    ws_s.cell(row=r, column=5).number_format = dol
    # 25yr NPV
    ws_s.cell(row=r, column=6, value=f'={sheet}!AC{net_r}').font = grn
    ws_s.cell(row=r, column=6).number_format = dol
    # Y1 net monthly
    ws_s.cell(row=r, column=7, value=f'={sheet}!C{net_r}/12').font = grn
    ws_s.cell(row=r, column=7).number_format = dol

# Battery incremental value section
sc(ws_s, 12, 1, "BATTERY INCREMENTAL ANALYSIS", sec)
sc(ws_s, 13, 1, "Utility", hdr, None, lbl)
sc(ws_s, 13, 2, "Battery Δ Cost ($)", hdr, None, lbl)
sc(ws_s, 13, 3, "Battery Δ Y1 Savings ($)", hdr, None, lbl)
sc(ws_s, 13, 4, "Battery Δ NPV ($)", hdr, None, lbl)

for i, (util, solar_row, batt_row) in enumerate([
    ("PG&E", 5, 6), ("SCE", 7, 8), ("SDG&E", 9, 10)
]):
    r = 14 + i
    sc(ws_s, r, 1, util, blk)
    ws_s.cell(row=r, column=2, value=f'=D{batt_row}-D{solar_row}').font = blk
    ws_s.cell(row=r, column=2).number_format = dol
    ws_s.cell(row=r, column=3, value=f'=E{batt_row}-E{solar_row}').font = blk
    ws_s.cell(row=r, column=3).number_format = dol
    ws_s.cell(row=r, column=4, value=f'=F{batt_row}-F{solar_row}').font = blk
    ws_s.cell(row=r, column=4).number_format = dol

print("Summary sheet done.")

# ================================================================
# SAVE
# ================================================================
out = HERE / "CA_Financial_Model_PhaseA.xlsx"
wb.save(out)
print(f"\nSaved: {out}")
print(f"Sheets: {wb.sheetnames}")
